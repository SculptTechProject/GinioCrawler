import csv
from datetime import datetime
import os, sys, re, asyncio, httpx, urllib.parse, urllib.robotparser as rp
from bs4 import BeautifulSoup as BS
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
PL_PHONE_RE = re.compile(r"(?:\+48\s?)?(?:\d{3}[\s-]?\d{3}[\s-]?\d{3})")


def write_excel(csv_path, xlsx_path):
    import pandas as pd
    from openpyxl.styles import Font

    df = pd.read_csv(csv_path, dtype=str, keep_default_na=False)

    for col in ("emails", "phones"):
        if col in df.columns:
            df[col] = (
                df[col]
                .str.replace(";", " ")
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Wyniki")
        ws = w.sheets["Wyniki"]
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(12, int(length * 0.9)), 60)
        cols = [c for c in ("url", "contact_url") if c in df.columns]
        for row in range(2, ws.max_row + 1):
            for name in cols:
                idx = df.columns.get_loc(name) + 1
                val = ws.cell(row, idx).value
                if isinstance(val, str) and val.startswith("http"):
                    cell = ws.cell(row, idx)
                    cell.hyperlink = val
                    cell.style = "Hyperlink"



def in_robots(url: str) -> bool:
    base = f"{urllib.parse.urlsplit(url).scheme}://{urllib.parse.urlsplit(url).netloc}"
    robots = rp.RobotFileParser()
    robots.set_url(urllib.parse.urljoin(base, "/robots.txt"))
    try:
        robots.read()
        return robots.can_fetch("*", url)
    except Exception:
        return True


def absolutize(base, href):
    return urllib.parse.urljoin(base, href)


async def fetch(url, client):
    try:
        if not in_robots(url):
            return None
        r = await client.get(url, timeout=15, follow_redirects=True)
        if r.status_code >= 400:
            return None
        return r.text
    except Exception:
        return None


def parse_info(html, url):
    soup = BS(html, "html.parser")
    title = (soup.title.string.strip() if soup.title else "")[:200]
    text = soup.get_text(" ", strip=True)
    emails = list(set(EMAIL_RE.findall(text)))
    phones = list(set(PL_PHONE_RE.findall(text)))
    contact = None
    for a in soup.select("a"):
        t = (a.get_text() or "").strip().lower()
        if t in ("kontakt", "contact") or "kontakt" in t or "contact" in t:
            contact = absolutize(url, a.get("href", ""))
            break
    return {
        "url": url,
        "title": title,
        "emails": emails,
        "phones": phones,
        "contact_url": contact,
    }


async def crawl_one(url, client):
    html = await fetch(url, client)
    if not html:
        return None
    base_info = parse_info(html, url)
    if base_info["contact_url"]:
        html2 = await fetch(base_info["contact_url"], client)
        if html2:
            extra = parse_info(html2, base_info["contact_url"])
            base_info["emails"] = list(set(base_info["emails"] + extra["emails"]))
            base_info["phones"] = list(set(base_info["phones"] + extra["phones"]))
    return base_info


async def search_serpapi(query, num=20):
    # wymaga SERPAPI_KEY w env
    key = os.environ["SERPAPI_KEY"]
    params = {"engine": "google", "q": query, "hl": "pl", "num": num, "api_key": key}
    async with httpx.AsyncClient() as client:
        r = await client.get("https://serpapi.com/search", params=params, timeout=20)
        data = r.json()
        urls = []
        for item in data.get("organic_results") or []:
            link = item.get("link")
            if link:
                urls.append(link)
        return urls


async def run(query):
    urls = await search_serpapi(query, num=25)
    async with httpx.AsyncClient(headers={"User-Agent": "Mozilla/5.0"}) as client:
        tasks = [crawl_one(u, client) for u in urls]
        results = [r for r in await asyncio.gather(*tasks) if r]
    # deduplikacja po emailach/URL
    seen = set()
    uniq = []
    for r in results:
        key = (tuple(sorted(r["emails"])), r["url"])
        if key in seen:
            continue
        seen.add(key)
        uniq.append(r)
    return uniq


if __name__ == "__main__":
    try:
        import os, sys, csv, asyncio
        from pathlib import Path
        from datetime import datetime

        q = input("Podaj szukane słowo: ").strip()
        if not q:
            print("Pusta fraza. Kończę.")
            sys.exit(1)

        if "SERPAPI_KEY" not in os.environ:
            print(
                "Brak SERPAPI_KEY w zmiennych środowiskowych. Ustaw i spróbuj ponownie."
            )
            sys.exit(2)

        data = asyncio.run(run(q))
        if not data:
            print("Brak wyników.")
            sys.exit(0)

        # katalogi wyjściowe
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_dir = Path("wyniki/csv")
        csv_dir.mkdir(parents=True, exist_ok=True)
        xlsx_dir = Path("wyniki/excel")
        xlsx_dir.mkdir(parents=True, exist_ok=True)

        outfile_csv = csv_dir / f"wyniki_{ts}.csv"
        outfile_xlsx = xlsx_dir / f"wyniki_{ts}.xlsx"

        # CSV (UTF-8-SIG ułatwia otwieranie w Excelu), separator = spacja
        with open(outfile_csv, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(
                f, fieldnames=["url", "title", "emails", "phones", "contact_url"]
            )
            w.writeheader()
            for row in data:
                row = {
                    **row,
                    "emails": " ".join(row.get("emails", [])),
                    "phones": " ".join(row.get("phones", [])),
                }
                w.writerow(row)

        print(f"Zapisano {len(data)} rekordów do: {outfile_csv}")

        # XLSX
        write_excel(str(outfile_csv), str(outfile_xlsx))
        print(f"Zapisano też Excel: {outfile_xlsx}")

    except KeyboardInterrupt:
        print("\nPrzerwano przez użytkownika.")
