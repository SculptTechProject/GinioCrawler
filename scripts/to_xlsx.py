import sys, pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def write_excel(csv_path, xlsx_path):
    df = pd.read_csv(csv_path)
    for col in ("emails", "phones"):
        if col in df:
            df[col] = df[col] = (
                df[col]
                .fillna("")
                .str.replace(";", " ")
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Wyniki")
        ws = w.sheets["Wyniki"]
        for c in ws[1]:
            c.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(
                max(12, int(length * 0.9)), 60
            )
        for row in range(2, ws.max_row + 1):
            for name in ("url", "contact_url"):
                if name in df:
                    idx = df.columns.get_loc(name) + 1
                    val = ws.cell(row, idx).value
                    if isinstance(val, str) and val.startswith("http"):
                        cell = ws.cell(row, idx)
                        cell.hyperlink = val
                        cell.style = "Hyperlink"


if __name__ == "__main__":
    csv_in = sys.argv[1] if len(sys.argv) > 1 else "wyniki.csv"
    xlsx_out = csv_in.replace(".csv", ".xlsx")
    write_excel(csv_in, xlsx_out)
    print(f"OK: {xlsx_out}")
